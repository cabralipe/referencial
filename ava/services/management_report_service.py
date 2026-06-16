from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Any
from xml.sax.saxutils import escape

from django.contrib.auth import get_user_model
from django.db.models import Count, Max, Q
from django.utils import timezone

from ava.models import (
    Atividade,
    AtividadeForumMensagem,
    AtividadeTentativa,
    Aula,
    ConteudoAula,
    Curso,
    CursoModulo,
    MatriculaCurso,
    ProgressoConteudo,
    ProgressoModulo,
)
from core.models import UserSessionLog
from curriculum.models import Escola, GT
from exports.services.pdf import html_to_pdf_bytes

try:  # pragma: no cover - dependencia validada pelo ambiente
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover - fallback defensivo
    Workbook = None  # type: ignore[assignment]
    Alignment = Font = PatternFill = None  # type: ignore[assignment]

try:  # pragma: no cover - dependencia validada pelo ambiente
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except ImportError:  # pragma: no cover - fallback defensivo
    REPORTLAB_AVAILABLE = False


class AVAManagementReportService:
    STATUS_LABELS = dict(AtividadeTentativa.Status.choices)
    TIPO_LABELS = dict(Atividade.Tipo.choices)

    @staticmethod
    def _is_super_admin(user) -> bool:
        return getattr(user, "role", None) == get_user_model().Role.SUPER_ADMIN

    @staticmethod
    def _display_user(usuario) -> str:
        return getattr(usuario, "nome", "") or usuario.get_full_name() or usuario.email

    @staticmethod
    def _to_local_datetime(value):
        if value is None:
            return None
        return timezone.localtime(value)

    @classmethod
    def _excel_datetime(cls, value):
        value = cls._to_local_datetime(value)
        if value is None:
            return None
        return value.replace(tzinfo=None)

    @classmethod
    def _merge_datetimes(cls, *values):
        validos = [cls._to_local_datetime(value) for value in values if value]
        if not validos:
            return None
        return max(validos)

    @classmethod
    def _scoped_user_queryset(cls, user):
        qs = get_user_model().objects.all()
        if not cls._is_super_admin(user):
            qs = qs.filter(cliente_id=user.cliente_id)
        return qs

    @classmethod
    def _scoped_school_queryset(cls, user):
        qs = Escola.objects.all()
        if not cls._is_super_admin(user):
            qs = qs.filter(cliente_id=user.cliente_id)
        return qs

    @classmethod
    def _scoped_course_queryset(cls, user):
        qs = Curso.objects.all()
        if not cls._is_super_admin(user):
            qs = qs.filter(cliente_id=user.cliente_id)
        return qs

    @classmethod
    def _scoped_module_queryset(cls, user):
        qs = CursoModulo.objects.select_related("curso")
        if not cls._is_super_admin(user):
            qs = qs.filter(curso__cliente_id=user.cliente_id)
        return qs

    @classmethod
    def _scoped_lesson_queryset(cls, user):
        qs = Aula.objects.select_related("modulo", "modulo__curso")
        if not cls._is_super_admin(user):
            qs = qs.filter(modulo__curso__cliente_id=user.cliente_id)
        return qs

    @classmethod
    def _resolve_course_ids(cls, user, filtros) -> list[int] | None:
        if filtros.get("curso_id"):
            return [filtros["curso_id"]]
        if filtros.get("modulo_id"):
            return list(
                cls._scoped_module_queryset(user)
                .filter(id=filtros["modulo_id"])
                .values_list("curso_id", flat=True)[:1]
            )
        if filtros.get("aula_id"):
            return list(
                cls._scoped_lesson_queryset(user)
                .filter(id=filtros["aula_id"])
                .values_list("modulo__curso_id", flat=True)[:1]
            )
        return None

    @staticmethod
    def _eixo_atividade_q(eixo_id):
        return (
            Q(atividade__eixos=eixo_id)
            | Q(atividade__aula__modulo__eixos=eixo_id)
            | Q(atividade__aula__modulo__curso__eixos=eixo_id)
        )

    @classmethod
    def _eixo_course_ids(cls, user, eixo_id):
        cursos = cls._scoped_course_queryset(user)
        ids = set(cursos.filter(eixos=eixo_id).values_list("id", flat=True))
        ids |= set(cursos.filter(modulos__eixos=eixo_id).values_list("id", flat=True))
        ids |= set(cursos.filter(modulos__aulas__atividades__eixos=eixo_id).values_list("id", flat=True))
        return ids

    @classmethod
    def _matriculas_queryset(cls, user, filtros):
        qs = MatriculaCurso.objects.select_related("aluno", "aluno__escola", "curso")
        if not cls._is_super_admin(user):
            qs = qs.filter(cliente_id=user.cliente_id)

        course_ids = cls._resolve_course_ids(user, filtros)
        if course_ids is not None:
            qs = qs.filter(curso_id__in=course_ids) if course_ids else qs.none()

        if filtros.get("usuario_id"):
            qs = qs.filter(aluno_id=filtros["usuario_id"])
        if filtros.get("escola_id"):
            qs = qs.filter(aluno__escola_id=filtros["escola_id"])
        if filtros.get("eixo_id"):
            qs = qs.filter(curso_id__in=cls._eixo_course_ids(user, filtros["eixo_id"]))

        if filtros.get("q"):
            termo = filtros["q"]
            qs = qs.filter(
                Q(aluno__nome__icontains=termo)
                | Q(aluno__email__icontains=termo)
                | Q(curso__titulo__icontains=termo)
            )

        return qs.order_by("curso__titulo", "aluno__nome", "aluno__email")

    @classmethod
    def _tentativas_queryset(cls, user, filtros):
        qs = AtividadeTentativa.objects.select_related(
            "aluno",
            "atividade__aula__modulo__curso",
        )
        if not cls._is_super_admin(user):
            qs = qs.filter(cliente_id=user.cliente_id)

        if filtros.get("usuario_id"):
            qs = qs.filter(aluno_id=filtros["usuario_id"])
        if filtros.get("escola_id"):
            qs = qs.filter(aluno__escola_id=filtros["escola_id"])
        if filtros.get("curso_id"):
            qs = qs.filter(atividade__aula__modulo__curso_id=filtros["curso_id"])
        if filtros.get("modulo_id"):
            qs = qs.filter(atividade__aula__modulo_id=filtros["modulo_id"])
        if filtros.get("aula_id"):
            qs = qs.filter(atividade__aula_id=filtros["aula_id"])
        if filtros.get("eixo_id"):
            qs = qs.filter(cls._eixo_atividade_q(filtros["eixo_id"])).distinct()
        if filtros.get("status"):
            qs = qs.filter(status=filtros["status"])
        if filtros.get("tipo"):
            qs = qs.filter(atividade__tipo=filtros["tipo"])
        if filtros.get("data_inicio"):
            qs = qs.filter(data_inicio__date__gte=filtros["data_inicio"])
        if filtros.get("data_fim"):
            qs = qs.filter(data_inicio__date__lte=filtros["data_fim"])
        if filtros.get("q"):
            termo = filtros["q"]
            qs = qs.filter(
                Q(aluno__nome__icontains=termo)
                | Q(aluno__email__icontains=termo)
                | Q(atividade__titulo__icontains=termo)
                | Q(atividade__aula__titulo__icontains=termo)
                | Q(atividade__aula__modulo__titulo__icontains=termo)
                | Q(atividade__aula__modulo__curso__titulo__icontains=termo)
            )
        return qs

    @classmethod
    def _forum_queryset(cls, user, filtros):
        if filtros.get("tipo") and filtros["tipo"] != Atividade.Tipo.FORUM:
            return AtividadeForumMensagem.objects.none()

        qs = AtividadeForumMensagem.objects.select_related(
            "autor",
            "atividade__aula__modulo__curso",
            "resposta_para",
        )
        if not cls._is_super_admin(user):
            qs = qs.filter(cliente_id=user.cliente_id)

        if filtros.get("usuario_id"):
            qs = qs.filter(autor_id=filtros["usuario_id"])
        if filtros.get("escola_id"):
            qs = qs.filter(autor__escola_id=filtros["escola_id"])
        if filtros.get("curso_id"):
            qs = qs.filter(atividade__aula__modulo__curso_id=filtros["curso_id"])
        if filtros.get("modulo_id"):
            qs = qs.filter(atividade__aula__modulo_id=filtros["modulo_id"])
        if filtros.get("aula_id"):
            qs = qs.filter(atividade__aula_id=filtros["aula_id"])
        if filtros.get("eixo_id"):
            qs = qs.filter(cls._eixo_atividade_q(filtros["eixo_id"])).distinct()
        if filtros.get("data_inicio"):
            qs = qs.filter(created_at__date__gte=filtros["data_inicio"])
        if filtros.get("data_fim"):
            qs = qs.filter(created_at__date__lte=filtros["data_fim"])
        if filtros.get("q"):
            termo = filtros["q"]
            qs = qs.filter(
                Q(autor__nome__icontains=termo)
                | Q(autor__email__icontains=termo)
                | Q(texto__icontains=termo)
                | Q(atividade__titulo__icontains=termo)
                | Q(atividade__aula__titulo__icontains=termo)
                | Q(atividade__aula__modulo__titulo__icontains=termo)
                | Q(atividade__aula__modulo__curso__titulo__icontains=termo)
            )
        return qs

    @classmethod
    def _content_queryset(cls, user, filtros):
        qs = ProgressoConteudo.objects.select_related(
            "progresso_aula__matricula__aluno",
            "conteudo__aula__modulo__curso",
        ).filter(is_visualizado=True)
        if not cls._is_super_admin(user):
            qs = qs.filter(cliente_id=user.cliente_id)

        if filtros.get("usuario_id"):
            qs = qs.filter(progresso_aula__matricula__aluno_id=filtros["usuario_id"])
        if filtros.get("escola_id"):
            qs = qs.filter(progresso_aula__matricula__aluno__escola_id=filtros["escola_id"])
        if filtros.get("curso_id"):
            qs = qs.filter(conteudo__aula__modulo__curso_id=filtros["curso_id"])
        if filtros.get("modulo_id"):
            qs = qs.filter(conteudo__aula__modulo_id=filtros["modulo_id"])
        if filtros.get("aula_id"):
            qs = qs.filter(conteudo__aula_id=filtros["aula_id"])
        if filtros.get("data_inicio"):
            qs = qs.filter(data_visualizacao__date__gte=filtros["data_inicio"])
        if filtros.get("data_fim"):
            qs = qs.filter(data_visualizacao__date__lte=filtros["data_fim"])
        if filtros.get("q"):
            termo = filtros["q"]
            qs = qs.filter(
                Q(progresso_aula__matricula__aluno__nome__icontains=termo)
                | Q(progresso_aula__matricula__aluno__email__icontains=termo)
                | Q(conteudo__titulo__icontains=termo)
                | Q(conteudo__aula__titulo__icontains=termo)
                | Q(conteudo__aula__modulo__titulo__icontains=termo)
                | Q(conteudo__aula__modulo__curso__titulo__icontains=termo)
            )
        return qs

    @classmethod
    def _selected_objects(cls, user, filtros) -> dict[str, Any]:
        usuario = None
        escola = None
        curso = None
        modulo = None
        aula = None

        if filtros.get("usuario_id"):
            usuario = cls._scoped_user_queryset(user).filter(id=filtros["usuario_id"]).first()
        if filtros.get("escola_id"):
            escola = cls._scoped_school_queryset(user).filter(id=filtros["escola_id"]).first()
        if filtros.get("curso_id"):
            curso = cls._scoped_course_queryset(user).filter(id=filtros["curso_id"]).first()
        if filtros.get("modulo_id"):
            modulo = cls._scoped_module_queryset(user).filter(id=filtros["modulo_id"]).first()
        if filtros.get("aula_id"):
            aula = cls._scoped_lesson_queryset(user).filter(id=filtros["aula_id"]).first()

        return {
            "usuario": usuario,
            "escola": escola,
            "curso": curso,
            "modulo": modulo,
            "aula": aula,
        }

    @classmethod
    def _filter_labels(cls, user, filtros, selected):
        labels = []
        if filtros.get("q"):
            labels.append(f'Busca: "{filtros["q"]}"')
        if selected["usuario"]:
            labels.append(f"Usuário: {cls._display_user(selected['usuario'])}")
        if selected["escola"]:
            labels.append(f"Escola: {selected['escola'].nome}")
        if selected["curso"]:
            labels.append(f"Curso: {selected['curso'].titulo}")
        if selected["modulo"]:
            labels.append(f"Módulo: {selected['modulo'].titulo}")
        if selected["aula"]:
            labels.append(f"Aula: {selected['aula'].titulo}")
        if filtros.get("status"):
            labels.append(f"Status da tentativa: {cls.STATUS_LABELS.get(filtros['status'], filtros['status'])}")
        if filtros.get("tipo"):
            labels.append(f"Tipo de atividade: {cls.TIPO_LABELS.get(filtros['tipo'], filtros['tipo'])}")
        if filtros.get("data_inicio"):
            labels.append(f"Interacoes a partir de: {filtros['data_inicio'].strftime('%d/%m/%Y')}")
        if filtros.get("data_fim"):
            labels.append(f"Interacoes ate: {filtros['data_fim'].strftime('%d/%m/%Y')}")
        if not labels:
            labels.append("Sem filtros adicionais. Relatório considera o escopo completo da gestão.")
        return labels

    @classmethod
    def build_report(cls, user, filtros) -> dict[str, Any]:
        selected = cls._selected_objects(user, filtros)
        matriculas = list(cls._matriculas_queryset(user, filtros))
        matricula_ids = [matricula.id for matricula in matriculas]
        curso_ids = sorted({matricula.curso_id for matricula in matriculas})
        aluno_ids = sorted({matricula.aluno_id for matricula in matriculas})
        matricula_by_key = {(matricula.aluno_id, matricula.curso_id): matricula for matricula in matriculas}

        gts_por_usuario: dict[int, list[str]] = defaultdict(list)
        if aluno_ids:
            gts_qs = GT.objects.filter(membros__id__in=aluno_ids)
            if not cls._is_super_admin(user):
                gts_qs = gts_qs.filter(cliente_id=user.cliente_id)
            for item in gts_qs.values("membros__id", "nome").order_by("nome"):
                gts_por_usuario[item["membros__id"]].append(item["nome"])

        acessos_por_usuario = {}
        if aluno_ids:
            session_qs = UserSessionLog.objects.filter(usuario_id__in=aluno_ids)
            if not cls._is_super_admin(user):
                session_qs = session_qs.filter(cliente_id=user.cliente_id)
            for item in session_qs.values("usuario_id").annotate(
                total_acessos=Count("id"),
                ultimo_acesso=Max("last_seen_at"),
            ):
                acessos_por_usuario[item["usuario_id"]] = item

        modulos_por_curso = defaultdict(list)
        if curso_ids:
            for modulo in CursoModulo.objects.filter(curso_id__in=curso_ids, is_active=True).order_by("curso_id", "ordem", "id"):
                modulos_por_curso[modulo.curso_id].append(modulo)

        progresso_modulos = {}
        if matricula_ids:
            progresso_modulos = {
                (progresso.matricula_id, progresso.modulo_id): progresso
                for progresso in ProgressoModulo.objects.filter(matricula_id__in=matricula_ids).select_related("modulo")
            }

        tentativas_stats = {}
        for item in cls._tentativas_queryset(user, filtros).values(
            "aluno_id",
            "atividade__aula__modulo__curso_id",
        ).annotate(
            envios_atividade=Count(
                "id",
                filter=Q(status__in=[AtividadeTentativa.Status.ENVIADA, AtividadeTentativa.Status.CORRIGIDA]),
                distinct=True,
            ),
            atividades_corrigidas=Count("id", filter=Q(status=AtividadeTentativa.Status.CORRIGIDA), distinct=True),
            envios_com_anexo=Count(
                "id",
                filter=Q(arquivos__isnull=False) | (Q(arquivo_enviado__isnull=False) & ~Q(arquivo_enviado="")),
                distinct=True,
            ),
            ultima_tentativa=Max("data_envio"),
        ):
            tentativas_stats[(item["aluno_id"], item["atividade__aula__modulo__curso_id"])] = item

        forum_stats = {}
        for item in cls._forum_queryset(user, filtros).values(
            "autor_id",
            "atividade__aula__modulo__curso_id",
        ).annotate(
            mensagens_forum=Count("id"),
            respostas_forum=Count("id", filter=Q(resposta_para__isnull=False)),
            ultima_mensagem=Max("created_at"),
        ):
            forum_stats[(item["autor_id"], item["atividade__aula__modulo__curso_id"])] = item

        conteudo_stats = {}
        for item in cls._content_queryset(user, filtros).values("progresso_aula__matricula_id").annotate(
            conteudos_visualizados=Count("id"),
            ultima_visualizacao=Max("data_visualizacao"),
        ):
            conteudo_stats[item["progresso_aula__matricula_id"]] = item

        linhas = []
        for matricula in matriculas:
            key = (matricula.aluno_id, matricula.curso_id)
            tentativas = tentativas_stats.get(key, {})
            forum = forum_stats.get(key, {})
            conteudos = conteudo_stats.get(matricula.id, {})

            concluidos = []
            pendentes = []
            for modulo in modulos_por_curso.get(matricula.curso_id, []):
                progresso = progresso_modulos.get((matricula.id, modulo.id))
                percentual = progresso.percentual if progresso else 0
                if progresso and progresso.is_concluido:
                    concluidos.append(modulo.titulo)
                else:
                    pendentes.append(modulo.titulo if percentual == 0 else f"{modulo.titulo} ({percentual}%)")

            ultima_interacao = cls._merge_datetimes(
                tentativas.get("ultima_tentativa"),
                forum.get("ultima_mensagem"),
                conteudos.get("ultima_visualizacao"),
            )

            linha = {
                "aluno_nome": cls._display_user(matricula.aluno),
                "aluno_email": matricula.aluno.email,
                "escola_nome": getattr(getattr(matricula.aluno, "escola", None), "nome", "") or "Sem escola vinculada",
                "gts": gts_por_usuario.get(matricula.aluno_id, []),
                "curso_titulo": matricula.curso.titulo,
                "status": matricula.status,
                "status_label": matricula.get_status_display(),
                "progresso_percentual": matricula.progresso_percentual,
                "modulos_total": len(modulos_por_curso.get(matricula.curso_id, [])),
                "modulos_concluidos": len(concluidos),
                "modulos_pendentes": len(pendentes),
                "lista_modulos_concluidos": concluidos,
                "lista_modulos_pendentes": pendentes,
                "envios_atividade": tentativas.get("envios_atividade", 0) or 0,
                "atividades_corrigidas": tentativas.get("atividades_corrigidas", 0) or 0,
                "envios_com_anexo": tentativas.get("envios_com_anexo", 0) or 0,
                "mensagens_forum": forum.get("mensagens_forum", 0) or 0,
                "respostas_forum": forum.get("respostas_forum", 0) or 0,
                "conteudos_visualizados": conteudos.get("conteudos_visualizados", 0) or 0,
                "total_acessos": acessos_por_usuario.get(matricula.aluno_id, {}).get("total_acessos", 0) or 0,
                "ultimo_acesso": cls._to_local_datetime(acessos_por_usuario.get(matricula.aluno_id, {}).get("ultimo_acesso")),
                "ultima_interacao": ultima_interacao,
                "data_matricula": cls._to_local_datetime(matricula.data_matricula),
                "data_conclusao": cls._to_local_datetime(matricula.data_conclusao),
            }
            linha["total_interacoes"] = (
                linha["envios_atividade"] + linha["mensagens_forum"] + linha["conteudos_visualizados"]
            )
            linhas.append(linha)

        top_interacao = sorted(
            linhas,
            key=lambda item: (
                item["total_interacoes"],
                item["mensagens_forum"],
                item["envios_atividade"],
                item["conteudos_visualizados"],
                item["progresso_percentual"],
            ),
            reverse=True,
        )
        pendentes = sorted(
            [linha for linha in linhas if linha["status"] != MatriculaCurso.Status.CONCLUIDA],
            key=lambda item: (
                item["progresso_percentual"],
                item["total_interacoes"],
                item["aluno_nome"].lower(),
            ),
        )

        total_alunos_unicos = len({linha["aluno_email"] for linha in linhas})
        concluidos_total = sum(1 for linha in linhas if linha["status"] == MatriculaCurso.Status.CONCLUIDA)
        progresso_medio = round(
            sum(linha["progresso_percentual"] for linha in linhas) / len(linhas),
            1,
        ) if linhas else 0.0
        resumo_por_escola = cls._resumir_linhas(linhas, "escola_nome")
        resumo_por_gt = cls._resumir_por_gt(linhas)

        escopo = "Todos os cursos do AVA"
        if selected["aula"]:
            escopo = f"Aula {selected['aula'].titulo}"
        elif selected["modulo"]:
            escopo = f"Módulo {selected['modulo'].titulo}"
        elif selected["curso"]:
            escopo = f"Curso {selected['curso'].titulo}"

        return {
            "generated_at": timezone.localtime(),
            "cliente_label": getattr(getattr(user, "cliente", None), "nome", "") or "Visão global",
            "escopo_label": escopo,
            "applied_filters": cls._filter_labels(user, filtros, selected),
            "linhas": linhas,
            "top_interacao": top_interacao,
            "pendentes": pendentes,
            "resumo_por_escola": resumo_por_escola,
            "resumo_por_gt": resumo_por_gt,
            "metricas": {
                "total_alunos_unicos": total_alunos_unicos,
                "total_matriculas": len(linhas),
                "concluidos": concluidos_total,
                "pendentes": len(linhas) - concluidos_total,
                "taxa_conclusao": round((concluidos_total / len(linhas)) * 100, 1) if linhas else 0.0,
                "progresso_medio": progresso_medio,
                "envios_atividade": sum(linha["envios_atividade"] for linha in linhas),
                "mensagens_forum": sum(linha["mensagens_forum"] for linha in linhas),
                "conteudos_visualizados": sum(linha["conteudos_visualizados"] for linha in linhas),
                "acessos_plataforma": sum(linha["total_acessos"] for linha in linhas),
                "sem_interacao": sum(1 for linha in linhas if linha["total_interacoes"] == 0),
                "ultima_interacao_geral": cls._merge_datetimes(
                    *(linha["ultima_interacao"] for linha in linhas if linha["ultima_interacao"])
                ),
            },
        }

    @classmethod
    def _resumir_linhas(cls, linhas: list[dict[str, Any]], campo: str) -> list[dict[str, Any]]:
        grupos: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for linha in linhas:
            grupos[linha.get(campo) or "-"].append(linha)
        return [cls._metricas_grupo(nome, itens) for nome, itens in sorted(grupos.items())]

    @classmethod
    def _resumir_por_gt(cls, linhas: list[dict[str, Any]]) -> list[dict[str, Any]]:
        grupos: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for linha in linhas:
            for nome in linha["gts"] or ["Sem GT vinculado"]:
                grupos[nome].append(linha)
        return [cls._metricas_grupo(nome, itens) for nome, itens in sorted(grupos.items())]

    @classmethod
    def _metricas_grupo(cls, nome: str, linhas: list[dict[str, Any]]) -> dict[str, Any]:
        concluidos = sum(1 for linha in linhas if linha["status"] == MatriculaCurso.Status.CONCLUIDA)
        progresso_medio = round(sum(linha["progresso_percentual"] for linha in linhas) / len(linhas), 1) if linhas else 0.0
        return {
            "nome": nome,
            "inscritos": len(linhas),
            "concluidos": concluidos,
            "pendentes": len(linhas) - concluidos,
            "taxa_conclusao": round((concluidos / len(linhas)) * 100, 1) if linhas else 0.0,
            "progresso_medio": progresso_medio,
            "envios_atividade": sum(linha["envios_atividade"] for linha in linhas),
            "atividades_corrigidas": sum(linha["atividades_corrigidas"] for linha in linhas),
            "acessos_plataforma": sum(linha["total_acessos"] for linha in linhas),
            "ultima_interacao": cls._merge_datetimes(*(linha["ultima_interacao"] for linha in linhas if linha["ultima_interacao"])),
        }

    @classmethod
    def render_pdf_bytes(cls, report_data: dict[str, Any]) -> bytes:
        if REPORTLAB_AVAILABLE:
            return cls._render_pdf_reportlab(report_data)

        context = {
            "titulo": "Relatório nominal de progresso e interações do AVA",
            **report_data,
            "top_interacao": report_data["top_interacao"][:10],
        }
        return html_to_pdf_bytes("ava/management/dashboard_report_pdf.html", context)

    @classmethod
    def _paragraph(cls, texto: str, style):
        return Paragraph(escape(texto or "-").replace("\n", "<br/>"), style)

    @classmethod
    def _modules_cell(cls, titulo: str, items: list[str], style):
        if not items:
            conteudo = f"<b>{escape(titulo)}:</b><br/>-"
        else:
            conteudo = f"<b>{escape(titulo)}:</b><br/>" + "<br/>".join(escape(item) for item in items)
        return Paragraph(conteudo, style)

    @classmethod
    def _datetime_text(cls, value) -> str:
        value = cls._to_local_datetime(value)
        return value.strftime("%d/%m/%Y %H:%M") if value else "-"

    @classmethod
    def _render_pdf_reportlab(cls, report_data: dict[str, Any]) -> bytes:
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=0.7 * cm,
            rightMargin=0.7 * cm,
            topMargin=1.1 * cm,
            bottomMargin=1.1 * cm,
            title="Relatório nominal de progresso e interações do AVA",
        )

        sample_styles = getSampleStyleSheet()
        styles = {
            "title": ParagraphStyle(
                "AVATitle",
                parent=sample_styles["Title"],
                fontSize=18,
                leading=22,
                textColor=colors.HexColor("#0F172A"),
                spaceAfter=6,
            ),
            "subtitle": ParagraphStyle(
                "AVASubtitle",
                parent=sample_styles["BodyText"],
                fontSize=9,
                leading=12,
                textColor=colors.HexColor("#475569"),
                spaceAfter=12,
            ),
            "section": ParagraphStyle(
                "AVASection",
                parent=sample_styles["Heading2"],
                fontSize=12,
                leading=15,
                textColor=colors.HexColor("#0F172A"),
                spaceBefore=8,
                spaceAfter=6,
            ),
            "body": ParagraphStyle(
                "AVABody",
                parent=sample_styles["BodyText"],
                fontSize=8,
                leading=10,
                textColor=colors.HexColor("#334155"),
            ),
            "small": ParagraphStyle(
                "AVASmall",
                parent=sample_styles["BodyText"],
                fontSize=7,
                leading=9,
                textColor=colors.HexColor("#334155"),
            ),
            "tiny": ParagraphStyle(
                "AVATiny",
                parent=sample_styles["BodyText"],
                fontSize=6.2,
                leading=7.4,
                textColor=colors.HexColor("#334155"),
                wordWrap="CJK",
            ),
            "metric": ParagraphStyle(
                "AVAMetric",
                parent=sample_styles["BodyText"],
                fontSize=10,
                leading=12,
                alignment=1,
                textColor=colors.HexColor("#0F172A"),
            ),
        }

        story = [
            cls._paragraph("Relatório nominal de progresso e interações do AVA", styles["title"]),
            cls._paragraph(
                "Snapshot atual das matrículas com contagem de envios e interações conforme os filtros aplicados na gestão.",
                styles["subtitle"],
            ),
        ]

        meta_table = Table(
            [
                [
                    cls._paragraph(f"Cliente: {report_data['cliente_label']}", styles["body"]),
                    cls._paragraph(f"Escopo: {report_data['escopo_label']}", styles["body"]),
                ],
                [
                    cls._paragraph(
                        f"Gerado em: {cls._datetime_text(report_data['generated_at'])}",
                        styles["body"],
                    ),
                    cls._paragraph(
                        f"Última interação geral: {cls._datetime_text(report_data['metricas']['ultima_interacao_geral'])}",
                        styles["body"],
                    ),
                ],
            ],
            colWidths=[13 * cm, 13 * cm],
        )
        meta_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.extend([meta_table, Spacer(1, 0.25 * cm)])

        story.append(cls._paragraph("Filtros aplicados", styles["section"]))
        for filtro in report_data["applied_filters"]:
            story.append(cls._paragraph(f"- {filtro}", styles["body"]))
        story.append(Spacer(1, 0.2 * cm))

        metricas = report_data["metricas"]
        metric_cells = [
            ("Alunos únicos", metricas["total_alunos_unicos"]),
            ("Matrículas", metricas["total_matriculas"]),
            ("Concluídas", metricas["concluidos"]),
            ("Pendentes", metricas["pendentes"]),
            ("Taxa de conclusão", f"{metricas['taxa_conclusao']}%"),
            ("Progresso médio", f"{metricas['progresso_medio']}%"),
            ("Envios", metricas["envios_atividade"]),
            ("Fórum", metricas["mensagens_forum"]),
            ("Conteúdos", metricas["conteudos_visualizados"]),
            ("Acessos", metricas["acessos_plataforma"]),
            ("Sem interação", metricas["sem_interacao"]),
        ]
        metric_rows = []
        for inicio in range(0, len(metric_cells), 5):
            row = []
            for label, value in metric_cells[inicio:inicio + 5]:
                row.append(
                    Paragraph(
                        f"<font size='7' color='#64748B'>{escape(label.upper())}</font><br/><b>{escape(str(value))}</b>",
                        styles["metric"],
                    )
                )
            metric_rows.append(row)

        metrics_table = Table(metric_rows, colWidths=[5.2 * cm] * 5)
        metrics_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.extend([cls._paragraph("Visão consolidada", styles["section"]), metrics_table, Spacer(1, 0.25 * cm)])

        for titulo, rows in (
            ("Consolidado por escola", report_data["resumo_por_escola"]),
            ("Consolidado por GT", report_data["resumo_por_gt"]),
        ):
            story.append(cls._paragraph(titulo, styles["section"]))
            resumo_table_data = [[
                cls._paragraph("Nome", styles["small"]),
                cls._paragraph("Inscritos", styles["small"]),
                cls._paragraph("Concl.", styles["small"]),
                cls._paragraph("Pend.", styles["small"]),
                cls._paragraph("Taxa", styles["small"]),
                cls._paragraph("Progresso", styles["small"]),
                cls._paragraph("Envios", styles["small"]),
                cls._paragraph("Corrig.", styles["small"]),
                cls._paragraph("Acessos", styles["small"]),
                cls._paragraph("Ultima interacao", styles["small"]),
            ]]
            for item in rows:
                resumo_table_data.append([
                    cls._paragraph(item["nome"], styles["body"]),
                    cls._paragraph(str(item["inscritos"]), styles["body"]),
                    cls._paragraph(str(item["concluidos"]), styles["body"]),
                    cls._paragraph(str(item["pendentes"]), styles["body"]),
                    cls._paragraph(f"{item['taxa_conclusao']}%", styles["body"]),
                    cls._paragraph(f"{item['progresso_medio']}%", styles["body"]),
                    cls._paragraph(str(item["envios_atividade"]), styles["body"]),
                    cls._paragraph(str(item["atividades_corrigidas"]), styles["body"]),
                    cls._paragraph(str(item["acessos_plataforma"]), styles["body"]),
                    cls._paragraph(cls._datetime_text(item["ultima_interacao"]), styles["body"]),
                ])
            resumo_table = Table(
                resumo_table_data,
                colWidths=[5.0 * cm, 1.8 * cm, 1.6 * cm, 1.6 * cm, 1.8 * cm, 2.1 * cm, 1.7 * cm, 1.7 * cm, 1.8 * cm, 4.0 * cm],
                repeatRows=1,
            )
            resumo_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ])
            )
            story.extend([resumo_table, Spacer(1, 0.18 * cm)])

        story.append(cls._paragraph("Top alunos que mais interagiram", styles["section"]))
        top_table_data = [[
            cls._paragraph("Aluno", styles["small"]),
            cls._paragraph("Curso", styles["small"]),
            cls._paragraph("Total", styles["small"]),
            cls._paragraph("Envios", styles["small"]),
            cls._paragraph("Fórum", styles["small"]),
            cls._paragraph("Conteúdos", styles["small"]),
            cls._paragraph("Última interação", styles["small"]),
        ]]
        top_rows = report_data["top_interacao"][:10] or []
        if top_rows:
            for linha in top_rows:
                top_table_data.append(
                    [
                        cls._paragraph(f"{linha['aluno_nome']}\n{linha['escola_nome']}\nGT: {', '.join(linha['gts']) or '-'}", styles["body"]),
                        cls._paragraph(linha["curso_titulo"], styles["body"]),
                        cls._paragraph(f"{linha['total_interacoes']}\n{linha['total_acessos']} acessos", styles["body"]),
                        cls._paragraph(str(linha["envios_atividade"]), styles["body"]),
                        cls._paragraph(str(linha["mensagens_forum"]), styles["body"]),
                        cls._paragraph(str(linha["conteudos_visualizados"]), styles["body"]),
                        cls._paragraph(cls._datetime_text(linha["ultima_interacao"]), styles["body"]),
                    ]
                )
        else:
            top_table_data.append([cls._paragraph("Nenhuma interação encontrada para o escopo selecionado.", styles["body"])] + [""] * 6)

        top_table = Table(top_table_data, colWidths=[4.2 * cm, 5.8 * cm, 2 * cm, 2 * cm, 2 * cm, 2.3 * cm, 3.5 * cm], repeatRows=1)
        top_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.extend([top_table, Spacer(1, 0.25 * cm)])

        story.append(cls._paragraph("Matrículas pendentes", styles["section"]))
        pendentes_table_data = [[
            cls._paragraph("Aluno", styles["small"]),
            cls._paragraph("Curso", styles["small"]),
            cls._paragraph("Status", styles["small"]),
            cls._paragraph("Progresso", styles["small"]),
            cls._paragraph("Módulos pendentes", styles["small"]),
            cls._paragraph("Última interação", styles["small"]),
        ]]
        if report_data["pendentes"]:
            for linha in report_data["pendentes"]:
                pendentes_table_data.append(
                    [
                        cls._paragraph(f"{linha['aluno_nome']}\n{linha['escola_nome']}\nGT: {', '.join(linha['gts']) or '-'}", styles["body"]),
                        cls._paragraph(linha["curso_titulo"], styles["body"]),
                        cls._paragraph(linha["status_label"], styles["body"]),
                        cls._paragraph(f"{linha['progresso_percentual']}%", styles["body"]),
                        cls._modules_cell("Pendentes", linha["lista_modulos_pendentes"], styles["small"]),
                        cls._paragraph(cls._datetime_text(linha["ultima_interacao"]), styles["body"]),
                    ]
                )
        else:
            pendentes_table_data.append([cls._paragraph("Não há matrículas pendentes neste recorte.", styles["body"])] + [""] * 5)

        pendentes_table = Table(
            pendentes_table_data,
            colWidths=[4.2 * cm, 5.8 * cm, 3 * cm, 2 * cm, 8 * cm, 3.2 * cm],
            repeatRows=1,
        )
        pendentes_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.extend([pendentes_table, PageBreak(), cls._paragraph("Detalhamento nominal dos alunos", styles["section"])])

        detalhe_table_data = [[
            cls._paragraph("Aluno", styles["small"]),
            cls._paragraph("Curso", styles["small"]),
            cls._paragraph("Status", styles["small"]),
            cls._paragraph("Progresso", styles["small"]),
            cls._paragraph("Módulos", styles["small"]),
            cls._paragraph("Envios", styles["small"]),
            cls._paragraph("Fórum", styles["small"]),
            cls._paragraph("Conteúdos", styles["small"]),
            cls._paragraph("Total", styles["small"]),
            cls._paragraph("Última interação", styles["small"]),
        ]]
        for linha in report_data["linhas"]:
            detalhe_table_data.append(
                [
                    cls._paragraph(f"{linha['aluno_nome']}\n{linha['aluno_email']}\n{linha['escola_nome']}\nGT: {', '.join(linha['gts']) or '-'}", styles["tiny"]),
                    cls._paragraph(linha["curso_titulo"], styles["tiny"]),
                    cls._paragraph(linha["status_label"], styles["tiny"]),
                    cls._paragraph(f"{linha['progresso_percentual']}%", styles["tiny"]),
                    Paragraph(
                        "<b>Concluídos:</b> "
                        + escape(", ".join(linha["lista_modulos_concluidos"]) or "-")
                        + "<br/><b>Pendentes:</b> "
                        + escape(", ".join(linha["lista_modulos_pendentes"]) or "-"),
                        styles["tiny"],
                    ),
                    cls._paragraph(
                        f"{linha['envios_atividade']} envios\n{linha['atividades_corrigidas']} corrigidas\n{linha['envios_com_anexo']} com anexo",
                        styles["tiny"],
                    ),
                    cls._paragraph(
                        f"{linha['mensagens_forum']} mensagens\n{linha['respostas_forum']} respostas",
                        styles["tiny"],
                    ),
                    cls._paragraph(f"{linha['conteudos_visualizados']} conteudos\n{linha['total_acessos']} acessos\nUlt. acesso: {cls._datetime_text(linha['ultimo_acesso'])}", styles["tiny"]),
                    cls._paragraph(str(linha["total_interacoes"]), styles["tiny"]),
                    cls._paragraph(cls._datetime_text(linha["ultima_interacao"]), styles["tiny"]),
                ]
            )

        detalhe_table = Table(
            detalhe_table_data,
            colWidths=[3.8 * cm, 4.0 * cm, 1.8 * cm, 1.5 * cm, 7.0 * cm, 2.1 * cm, 2.0 * cm, 1.6 * cm, 1.3 * cm, 3.0 * cm],
            repeatRows=1,
        )
        detalhe_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(detalhe_table)

        document.build(story)
        buffer.seek(0)
        return buffer.read()

    @classmethod
    def render_xlsx_bytes(cls, report_data: dict[str, Any]) -> bytes:
        if Workbook is None:
            raise RuntimeError("openpyxl não está disponível no ambiente.")

        workbook = Workbook()
        resumo = workbook.active
        resumo.title = "Resumo"

        cabecalho_fill = PatternFill("solid", fgColor="1D4ED8")
        cabecalho_font = Font(color="FFFFFF", bold=True)
        titulo_font = Font(size=16, bold=True)
        destaque_font = Font(size=12, bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        resumo["A1"] = "Relatório nominal de progresso e interações do AVA"
        resumo["A1"].font = titulo_font
        resumo["A3"] = "Gerado em"
        resumo["B3"] = cls._excel_datetime(report_data["generated_at"])
        resumo["B3"].number_format = "dd/mm/yyyy hh:mm"
        resumo["A4"] = "Cliente"
        resumo["B4"] = report_data["cliente_label"]
        resumo["A5"] = "Escopo"
        resumo["B5"] = report_data["escopo_label"]

        resumo["A7"] = "Filtros aplicados"
        resumo["A7"].font = destaque_font
        for indice, filtro in enumerate(report_data["applied_filters"], start=8):
            resumo[f"A{indice}"] = filtro

        metricas_inicio = 8 + len(report_data["applied_filters"]) + 2
        resumo[f"A{metricas_inicio}"] = "Indicador"
        resumo[f"B{metricas_inicio}"] = "Valor"
        for coluna in ("A", "B"):
            resumo[f"{coluna}{metricas_inicio}"].fill = cabecalho_fill
            resumo[f"{coluna}{metricas_inicio}"].font = cabecalho_font

        metricas_labels = [
            ("Alunos únicos", report_data["metricas"]["total_alunos_unicos"]),
            ("Matrículas no relatório", report_data["metricas"]["total_matriculas"]),
            ("Matrículas concluídas", report_data["metricas"]["concluidos"]),
            ("Matrículas pendentes", report_data["metricas"]["pendentes"]),
            ("Taxa de conclusão (%)", report_data["metricas"]["taxa_conclusao"]),
            ("Progresso médio (%)", report_data["metricas"]["progresso_medio"]),
            ("Envios de atividade", report_data["metricas"]["envios_atividade"]),
            ("Mensagens no fórum", report_data["metricas"]["mensagens_forum"]),
            ("Conteúdos visualizados", report_data["metricas"]["conteudos_visualizados"]),
            ("Alunos sem interação", report_data["metricas"]["sem_interacao"]),
        ]
        for linha_indice, (label, valor) in enumerate(metricas_labels, start=metricas_inicio + 1):
            resumo[f"A{linha_indice}"] = label
            resumo[f"B{linha_indice}"] = valor

        resumo.column_dimensions["A"].width = 28
        resumo.column_dimensions["B"].width = 24

        alunos_sheet = workbook.create_sheet("Alunos")
        headers = [
            "Aluno",
            "E-mail",
            "Escola",
            "GTs",
            "Curso",
            "Status da matricula",
            "Progresso (%)",
            "Modulos concluidos",
            "Modulos pendentes",
            "Lista de modulos concluidos",
            "Lista de modulos pendentes",
            "Envios de atividade",
            "Atividades corrigidas",
            "Envios com anexo",
            "Mensagens no forum",
            "Respostas no forum",
            "Conteudos visualizados",
            "Acessos a plataforma",
            "Ultimo acesso",
            "Total de interacoes",
            "Ultima interacao",
            "Data da matricula",
            "Data de conclusao",
        ]
        alunos_sheet.append(headers)
        for cell in alunos_sheet[1]:
            cell.fill = cabecalho_fill
            cell.font = cabecalho_font

        for linha in report_data["linhas"]:
            alunos_sheet.append(
                [
                    linha["aluno_nome"],
                    linha["aluno_email"],
                    linha["escola_nome"],
                    ", ".join(linha["gts"]) or "-",
                    linha["curso_titulo"],
                    linha["status_label"],
                    linha["progresso_percentual"],
                    linha["modulos_concluidos"],
                    linha["modulos_pendentes"],
                    ", ".join(linha["lista_modulos_concluidos"]) or "-",
                    ", ".join(linha["lista_modulos_pendentes"]) or "-",
                    linha["envios_atividade"],
                    linha["atividades_corrigidas"],
                    linha["envios_com_anexo"],
                    linha["mensagens_forum"],
                    linha["respostas_forum"],
                    linha["conteudos_visualizados"],
                    linha["total_acessos"],
                    cls._excel_datetime(linha["ultimo_acesso"]),
                    linha["total_interacoes"],
                    cls._excel_datetime(linha["ultima_interacao"]),
                    cls._excel_datetime(linha["data_matricula"]),
                    cls._excel_datetime(linha["data_conclusao"]),
                ]
            )

        for row in alunos_sheet.iter_rows(min_row=2):
            row[3].alignment = wrap_alignment
            row[9].alignment = wrap_alignment
            row[10].alignment = wrap_alignment
            for idx in (18, 20, 21, 22):
                if row[idx].value:
                    row[idx].number_format = "dd/mm/yyyy hh:mm"

        alunos_sheet.freeze_panes = "A2"
        alunos_sheet.auto_filter.ref = alunos_sheet.dimensions
        larguras = {
            "A": 24,
            "B": 28,
            "C": 28,
            "D": 20,
            "E": 14,
            "F": 18,
            "G": 18,
            "H": 40,
            "I": 40,
            "J": 18,
            "K": 18,
            "L": 18,
            "M": 18,
            "N": 18,
            "O": 20,
            "P": 18,
            "Q": 18,
            "R": 18,
            "S": 18,
        }
        for coluna, largura in larguras.items():
            alunos_sheet.column_dimensions[coluna].width = largura

        top_sheet = workbook.create_sheet("Top interação")
        top_headers = [
            "Aluno",
            "Curso",
            "Total de interações",
            "Envios de atividade",
            "Mensagens no fórum",
            "Conteúdos visualizados",
            "Última interação",
        ]
        top_sheet.append(top_headers)
        for cell in top_sheet[1]:
            cell.fill = cabecalho_fill
            cell.font = cabecalho_font
        for linha in report_data["top_interacao"]:
            top_sheet.append(
                [
                    linha["aluno_nome"],
                    linha["curso_titulo"],
                    linha["total_interacoes"],
                    linha["envios_atividade"],
                    linha["mensagens_forum"],
                    linha["conteudos_visualizados"],
                    cls._excel_datetime(linha["ultima_interacao"]),
                ]
            )
        for row in top_sheet.iter_rows(min_row=2):
            if row[6].value:
                row[6].number_format = "dd/mm/yyyy hh:mm"
        for coluna, largura in {"A": 24, "B": 28, "C": 18, "D": 18, "E": 18, "F": 20, "G": 18}.items():
            top_sheet.column_dimensions[coluna].width = largura

        pendentes_sheet = workbook.create_sheet("Pendentes")
        pendentes_headers = [
            "Aluno",
            "Curso",
            "Status da matrícula",
            "Progresso (%)",
            "Módulos pendentes",
            "Lista de módulos pendentes",
            "Envios de atividade",
            "Mensagens no fórum",
            "Última interação",
        ]
        pendentes_sheet.append(pendentes_headers)
        for cell in pendentes_sheet[1]:
            cell.fill = cabecalho_fill
            cell.font = cabecalho_font
        for linha in report_data["pendentes"]:
            pendentes_sheet.append(
                [
                    linha["aluno_nome"],
                    linha["curso_titulo"],
                    linha["status_label"],
                    linha["progresso_percentual"],
                    linha["modulos_pendentes"],
                    ", ".join(linha["lista_modulos_pendentes"]) or "-",
                    linha["envios_atividade"],
                    linha["mensagens_forum"],
                    cls._excel_datetime(linha["ultima_interacao"]),
                ]
            )
        for row in pendentes_sheet.iter_rows(min_row=2):
            row[5].alignment = wrap_alignment
            if row[8].value:
                row[8].number_format = "dd/mm/yyyy hh:mm"
        for coluna, largura in {"A": 24, "B": 28, "C": 20, "D": 14, "E": 18, "F": 40, "G": 18, "H": 18, "I": 18}.items():
            pendentes_sheet.column_dimensions[coluna].width = largura

        for sheet_name, rows in (
            ("Por escola", report_data["resumo_por_escola"]),
            ("Por GT", report_data["resumo_por_gt"]),
        ):
            sheet = workbook.create_sheet(sheet_name)
            summary_headers = [
                "Nome",
                "Inscritos",
                "Concluidos",
                "Pendentes",
                "Taxa de conclusao (%)",
                "Progresso medio (%)",
                "Envios de atividade",
                "Atividades corrigidas",
                "Acessos a plataforma",
                "Ultima interacao",
            ]
            sheet.append(summary_headers)
            for cell in sheet[1]:
                cell.fill = cabecalho_fill
                cell.font = cabecalho_font
            for item in rows:
                sheet.append(
                    [
                        item["nome"],
                        item["inscritos"],
                        item["concluidos"],
                        item["pendentes"],
                        item["taxa_conclusao"],
                        item["progresso_medio"],
                        item["envios_atividade"],
                        item["atividades_corrigidas"],
                        item["acessos_plataforma"],
                        cls._excel_datetime(item["ultima_interacao"]),
                    ]
                )
            for row in sheet.iter_rows(min_row=2):
                if row[9].value:
                    row[9].number_format = "dd/mm/yyyy hh:mm"
            for coluna, largura in {"A": 30, "B": 14, "C": 14, "D": 14, "E": 20, "F": 20, "G": 18, "H": 20, "I": 20, "J": 18}.items():
                sheet.column_dimensions[coluna].width = largura

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()
