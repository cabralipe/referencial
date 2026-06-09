from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from ava.models import (
    Atividade,
    AtividadeForumMensagem,
    AtividadeTentativa,
    AtividadeTentativaArquivo,
    Aula,
    ConteudoAula,
    Curso,
    CursoModulo,
    MatriculaCurso,
    ProgressoAula,
    ProgressoConteudo,
    ProgressoModulo,
    QuizAlternativa,
    QuizQuestao,
    QuizRespostaItem,
)
from core.models import Cliente
from curriculum.models import Escola


User = get_user_model()


class AVAManagementDashboardTests(TestCase):
    def setUp(self):
        self.cliente = Cliente.objects.create(nome="Cliente Gestao", slug="cliente-gestao")
        self.escola_1 = Escola.objects.create(cliente=self.cliente, nome="Escola Alpha")
        self.escola_2 = Escola.objects.create(cliente=self.cliente, nome="Escola Beta")
        self.admin = User.objects.create_user(
            email="admin@gestao.com",
            nome="Admin Gestao",
            password="123456",
            cliente=self.cliente,
            role=User.Role.ADMIN_CLIENTE,
        )
        self.redator = User.objects.create_user(
            email="redator@gestao.com",
            nome="Redator Gestao",
            password="123456",
            cliente=self.cliente,
            role=User.Role.ARTICULADOR,
        )
        self.revisor = User.objects.create_user(
            email="revisor@gestao.com",
            nome="Revisor Gestao",
            password="123456",
            cliente=self.cliente,
            role=User.Role.REVISOR,
        )
        self.leitor = User.objects.create_user(
            email="leitor@gestao.com",
            nome="Leitor Gestao",
            password="123456",
            cliente=self.cliente,
            role=User.Role.LEITOR,
        )
        self.aluno_1 = User.objects.create_user(
            email="aluno1@gestao.com",
            nome="Aluno Um",
            password="123456",
            cliente=self.cliente,
            escola=self.escola_1,
            role=User.Role.LEITOR,
        )
        self.aluno_2 = User.objects.create_user(
            email="aluno2@gestao.com",
            nome="Aluno Dois",
            password="123456",
            cliente=self.cliente,
            escola=self.escola_2,
            role=User.Role.LEITOR,
        )

        self.curso = Curso.objects.create(
            cliente=self.cliente,
            titulo="Curso Gestao AVA",
            slug="curso-gestao-ava",
            status=Curso.Status.PUBLICADO,
            is_aberto=True,
            autor_principal=self.admin,
        )
        self.modulo = CursoModulo.objects.create(
            cliente=self.cliente,
            curso=self.curso,
            titulo="Modulo Gestao",
            ordem=1,
            is_active=True,
        )
        self.aula = Aula.objects.create(
            cliente=self.cliente,
            modulo=self.modulo,
            titulo="Aula Gestao",
            ordem=1,
            is_active=True,
        )
        self.modulo_2 = CursoModulo.objects.create(
            cliente=self.cliente,
            curso=self.curso,
            titulo="Modulo Complementar",
            ordem=2,
            is_active=True,
        )
        self.aula_2 = Aula.objects.create(
            cliente=self.cliente,
            modulo=self.modulo_2,
            titulo="Aula Complementar",
            ordem=1,
            is_active=True,
        )
        self.atividade_quiz = Atividade.objects.create(
            cliente=self.cliente,
            aula=self.aula,
            tipo=Atividade.Tipo.QUIZ,
            titulo="Quiz Gestao",
        )
        self.atividade_tarefa = Atividade.objects.create(
            cliente=self.cliente,
            aula=self.aula,
            tipo=Atividade.Tipo.TAREFA,
            titulo="Tarefa Gestao",
        )
        self.atividade_forum = Atividade.objects.create(
            cliente=self.cliente,
            aula=self.aula_2,
            tipo=Atividade.Tipo.FORUM,
            titulo="Forum Gestao",
        )
        self.conteudo_1 = ConteudoAula.objects.create(
            cliente=self.cliente,
            aula=self.aula,
            tipo=ConteudoAula.Tipo.TEXTO,
            titulo="Conteudo base",
            conteudo_texto="Texto principal",
            ordem=1,
        )
        self.conteudo_2 = ConteudoAula.objects.create(
            cliente=self.cliente,
            aula=self.aula_2,
            tipo=ConteudoAula.Tipo.TEXTO,
            titulo="Conteudo complementar",
            conteudo_texto="Texto complementar",
            ordem=1,
        )

        self.questao = QuizQuestao.objects.create(
            cliente=self.cliente,
            atividade=self.atividade_quiz,
            enunciado="Questao da gestao?",
            ordem=1,
        )
        self.alt_correta = QuizAlternativa.objects.create(
            cliente=self.cliente,
            questao=self.questao,
            texto="Alternativa correta",
            is_correta=True,
            ordem=1,
        )
        self.alt_errada = QuizAlternativa.objects.create(
            cliente=self.cliente,
            questao=self.questao,
            texto="Alternativa errada",
            is_correta=False,
            ordem=2,
        )

        self.matricula_1 = MatriculaCurso.objects.create(
            cliente=self.cliente,
            curso=self.curso,
            aluno=self.aluno_1,
            matriculado_por=self.admin,
        )
        self.matricula_2 = MatriculaCurso.objects.create(
            cliente=self.cliente,
            curso=self.curso,
            aluno=self.aluno_2,
            matriculado_por=self.admin,
        )

        self.tentativa_1 = AtividadeTentativa.objects.create(
            cliente=self.cliente,
            atividade=self.atividade_quiz,
            aluno=self.aluno_1,
            status=AtividadeTentativa.Status.CORRIGIDA,
            nota_obtida=100,
        )
        self.tentativa_2 = AtividadeTentativa.objects.create(
            cliente=self.cliente,
            atividade=self.atividade_tarefa,
            aluno=self.aluno_2,
            status=AtividadeTentativa.Status.ENVIADA,
            texto_resposta="Minha resposta discursiva",
        )
        AtividadeTentativaArquivo.objects.create(
            cliente=self.cliente,
            tentativa=self.tentativa_2,
            arquivo=SimpleUploadedFile("atividade-gestao.pdf", b"%PDF-1.4", content_type="application/pdf"),
            nome_original="atividade-gestao.pdf",
        )
        QuizRespostaItem.objects.create(
            cliente=self.cliente,
            tentativa=self.tentativa_1,
            questao=self.questao,
            alternativa_selecionada=self.alt_correta,
            is_correta=True,
            nota_item=100,
        )

        agora = timezone.now()
        progresso_aula_1 = ProgressoAula.objects.create(
            cliente=self.cliente,
            matricula=self.matricula_1,
            aula=self.aula,
            is_concluida=True,
            data_conclusao=agora,
        )
        progresso_aula_2 = ProgressoAula.objects.create(
            cliente=self.cliente,
            matricula=self.matricula_1,
            aula=self.aula_2,
            is_concluida=False,
        )
        ProgressoConteudo.objects.create(
            cliente=self.cliente,
            progresso_aula=progresso_aula_1,
            conteudo=self.conteudo_1,
            is_visualizado=True,
            data_visualizacao=agora,
        )
        ProgressoConteudo.objects.create(
            cliente=self.cliente,
            progresso_aula=progresso_aula_2,
            conteudo=self.conteudo_2,
            is_visualizado=True,
            data_visualizacao=agora,
        )
        ProgressoModulo.objects.create(
            cliente=self.cliente,
            matricula=self.matricula_1,
            modulo=self.modulo,
            is_concluido=True,
            percentual=100,
            data_conclusao=agora,
        )
        ProgressoModulo.objects.create(
            cliente=self.cliente,
            matricula=self.matricula_1,
            modulo=self.modulo_2,
            is_concluido=False,
            percentual=50,
        )
        ProgressoModulo.objects.create(
            cliente=self.cliente,
            matricula=self.matricula_2,
            modulo=self.modulo,
            is_concluido=False,
            percentual=0,
        )
        ProgressoAula.objects.create(
            cliente=self.cliente,
            matricula=self.matricula_2,
            aula=self.aula,
            is_concluida=False,
        )
        self.matricula_1.progresso_percentual = 50
        self.matricula_1.save(update_fields=["progresso_percentual"])

        self.forum_msg_1 = AtividadeForumMensagem.objects.create(
            cliente=self.cliente,
            atividade=self.atividade_forum,
            autor=self.aluno_1,
            texto="Primeira interacao no forum",
        )
        AtividadeForumMensagem.objects.create(
            cliente=self.cliente,
            atividade=self.atividade_forum,
            autor=self.aluno_1,
            texto="Resposta em thread",
            resposta_para=self.forum_msg_1,
        )
        AtividadeForumMensagem.objects.create(
            cliente=self.cliente,
            atividade=self.atividade_forum,
            autor=self.aluno_2,
            texto="Participacao do aluno dois",
        )

        self.outro_cliente = Cliente.objects.create(nome="Cliente Externo", slug="cliente-externo")
        self.aluno_externo = User.objects.create_user(
            email="aluno@externo.com",
            nome="Aluno Externo",
            password="123456",
            cliente=self.outro_cliente,
            role=User.Role.LEITOR,
        )
        self.curso_externo = Curso.objects.create(
            cliente=self.outro_cliente,
            titulo="Curso Externo",
            slug="curso-externo",
            status=Curso.Status.PUBLICADO,
            is_aberto=True,
            autor_principal=self.aluno_externo,
        )
        modulo_externo = CursoModulo.objects.create(
            cliente=self.outro_cliente,
            curso=self.curso_externo,
            titulo="Modulo Externo",
            ordem=1,
            is_active=True,
        )
        aula_externa = Aula.objects.create(
            cliente=self.outro_cliente,
            modulo=modulo_externo,
            titulo="Aula Externa",
            ordem=1,
            is_active=True,
        )
        atividade_externa = Atividade.objects.create(
            cliente=self.outro_cliente,
            aula=aula_externa,
            tipo=Atividade.Tipo.TAREFA,
            titulo="Tarefa Externa",
        )
        MatriculaCurso.objects.create(
            cliente=self.outro_cliente,
            curso=self.curso_externo,
            aluno=self.aluno_externo,
            matriculado_por=self.aluno_externo,
        )
        self.tentativa_externa = AtividadeTentativa.objects.create(
            cliente=self.outro_cliente,
            atividade=atividade_externa,
            aluno=self.aluno_externo,
            status=AtividadeTentativa.Status.ENVIADA,
            texto_resposta="Resposta externa",
        )

    def test_dashboard_requires_admin_or_redator_role(self):
        self.client.force_login(self.leitor)
        response = self.client.get(reverse("ava:gestao_dashboard"))
        self.assertEqual(response.status_code, 403)

    def test_dashboard_allows_admin_redator_and_revisor(self):
        self.client.force_login(self.admin)
        response_admin = self.client.get(reverse("ava:gestao_dashboard"))
        self.assertEqual(response_admin.status_code, 200)
        self.assertContains(response_admin, "Dashboard de respostas e utilizacao")

        self.client.force_login(self.redator)
        response_redator = self.client.get(reverse("ava:gestao_dashboard"))
        self.assertEqual(response_redator.status_code, 200)
        self.assertContains(response_redator, "Tentativas e respostas dos usuarios")

        self.client.force_login(self.revisor)
        response_revisor = self.client.get(reverse("ava:gestao_dashboard"))
        self.assertEqual(response_revisor.status_code, 200)
        self.assertContains(response_revisor, "Tentativas e respostas dos usuarios")

    def test_dashboard_exibe_acesso_a_pagina_explicativa_do_relatorio(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("ava:gestao_dashboard_relatorio_entenda"))

    def test_dashboard_relatorio_entenda_exige_permissao_de_gestao(self):
        self.client.force_login(self.leitor)
        response = self.client.get(reverse("ava:gestao_dashboard_relatorio_entenda"))

        self.assertEqual(response.status_code, 403)

    def test_dashboard_relatorio_entenda_renderiza_guia(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_dashboard_relatorio_entenda"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "PDF")
        self.assertContains(response, reverse("ava:gestao_dashboard"))

    def test_dashboard_can_filter_by_usuario(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_dashboard"), {"usuario": str(self.aluno_1.id)})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["metricas"]["total_tentativas"], 1)
        tentativa_ids = {tentativa.id for tentativa in response.context["page_obj"].object_list}
        self.assertEqual(tentativa_ids, {self.tentativa_1.id})

    def test_dashboard_can_filter_by_escola(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_dashboard"), {"escola": str(self.escola_1.id)})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["filtros"]["escola_id"], self.escola_1.id)
        self.assertEqual(response.context["visao_geral"]["total_alunos"], 1)
        self.assertEqual(response.context["metricas"]["total_tentativas"], 1)
        tentativa_ids = {tentativa.id for tentativa in response.context["page_obj"].object_list}
        self.assertEqual(tentativa_ids, {self.tentativa_1.id})

    def test_dashboard_exibe_visao_geral_real_do_cliente(self):
        curso_sem_tentativa = Curso.objects.create(
            cliente=self.cliente,
            titulo="Curso Sem Tentativa",
            slug="curso-sem-tentativa",
            status=Curso.Status.PUBLICADO,
            is_aberto=True,
            autor_principal=self.admin,
        )
        MatriculaCurso.objects.create(
            cliente=self.cliente,
            curso=curso_sem_tentativa,
            aluno=self.aluno_1,
            matriculado_por=self.admin,
        )

        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["visao_geral"]["total_alunos"], 2)
        self.assertEqual(response.context["visao_geral"]["total_matriculas"], 3)
        self.assertEqual(response.context["visao_geral"]["total_cursos"], 2)
        self.assertEqual(response.context["visao_geral"]["total_atividades"], 3)
        curso_ids = {curso.id for curso in response.context["cursos"]}
        self.assertIn(curso_sem_tentativa.id, curso_ids)

    def test_dashboard_restringe_metricas_ao_cliente_do_admin(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["metricas"]["total_tentativas"], 2)
        tentativa_ids = {tentativa.id for tentativa in response.context["page_obj"].object_list}
        self.assertNotIn(self.tentativa_externa.id, tentativa_ids)
        curso_ids = {curso.id for curso in response.context["cursos"]}
        self.assertNotIn(self.curso_externo.id, curso_ids)

    def test_tentativa_detalhe_shows_quiz_resposta(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_tentativa_detalhe", args=[self.tentativa_1.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Questao da gestao?")
        self.assertContains(response, "Alternativa correta")
        self.assertContains(response, "Salvar correcao")

    def test_tentativa_detalhe_shows_uploaded_attempt_files(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_tentativa_detalhe", args=[self.tentativa_2.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Anexos enviados:")
        self.assertContains(response, "atividade-gestao.pdf")

    def test_tentativa_detalhe_permite_corrigir_tentativa_pelo_ava(self):
        self.client.force_login(self.revisor)
        response = self.client.post(
            reverse("ava:gestao_tentativa_detalhe", args=[self.tentativa_2.id]),
            {
                "status": AtividadeTentativa.Status.CORRIGIDA,
                "nota_obtida": "87.50",
                "feedback_tutor": "Resposta consistente com os criterios do modulo.",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Correcao salva com sucesso.")

        self.tentativa_2.refresh_from_db()
        self.assertEqual(self.tentativa_2.status, AtividadeTentativa.Status.CORRIGIDA)
        self.assertEqual(str(self.tentativa_2.nota_obtida), "87.50")
        self.assertEqual(self.tentativa_2.feedback_tutor, "Resposta consistente com os criterios do modulo.")
        self.assertEqual(self.tentativa_2.corrigido_por, self.revisor)
        self.assertIsNotNone(self.tentativa_2.data_correcao)

    def test_tentativa_detalhe_bloqueia_acesso_a_dado_de_outro_cliente(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_tentativa_detalhe", args=[self.tentativa_externa.id]))
        self.assertEqual(response.status_code, 404)

    def test_dashboard_relatorio_xlsx_exporta_dados_nominais(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ava:gestao_dashboard_relatorio", args=["xlsx"]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("relatorio-ava-nominal-", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Resumo", "Alunos", "Top interacao", "Pendentes"])

        alunos_sheet = workbook["Alunos"]
        alunos = {
            row[0]: row
            for row in alunos_sheet.iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        self.assertIn("Aluno Um", alunos)
        self.assertIn("Aluno Dois", alunos)
        self.assertNotIn("Aluno Externo", alunos)

        aluno_um = alunos["Aluno Um"]
        self.assertEqual(aluno_um[2], "Escola Alpha")
        self.assertEqual(aluno_um[4], "Curso Gestao AVA")
        self.assertEqual(aluno_um[6], 50)
        self.assertEqual(aluno_um[7], 1)
        self.assertEqual(aluno_um[8], 1)
        self.assertIn("Modulo Gestao", aluno_um[9])
        self.assertIn("Modulo Complementar (50%)", aluno_um[10])
        self.assertEqual(aluno_um[11], 1)
        self.assertEqual(aluno_um[12], 1)
        self.assertEqual(aluno_um[14], 2)
        self.assertEqual(aluno_um[15], 1)
        self.assertEqual(aluno_um[16], 2)
        self.assertEqual(aluno_um[19], 5)

    def test_dashboard_relatorio_xlsx_respeita_filtro_de_usuario(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("ava:gestao_dashboard_relatorio", args=["xlsx"]),
            {"usuario": str(self.aluno_1.id)},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        alunos_sheet = workbook["Alunos"]
        alunos = [row[0] for row in alunos_sheet.iter_rows(min_row=2, values_only=True) if row[0]]
        self.assertEqual(alunos, ["Aluno Um"])

    def test_dashboard_relatorio_xlsx_respeita_filtro_de_escola(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("ava:gestao_dashboard_relatorio", args=["xlsx"]),
            {"escola": str(self.escola_1.id)},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        resumo_values = [cell.value for cell in workbook["Resumo"]["A"] if cell.value]
        alunos = [row[0] for row in workbook["Alunos"].iter_rows(min_row=2, values_only=True) if row[0]]

        self.assertIn("Escola: Escola Alpha", resumo_values)
        self.assertEqual(alunos, ["Aluno Um"])

    def test_dashboard_relatorio_pdf_retorna_arquivo_valido(self):
        self.client.force_login(self.revisor)
        response = self.client.get(reverse("ava:gestao_dashboard_relatorio", args=["pdf"]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("relatorio-ava-nominal-", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))
